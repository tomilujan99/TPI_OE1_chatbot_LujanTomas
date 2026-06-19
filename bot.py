"""
Bot de consulta y pedido de repuestos - Simulacion WhatsApp Business
Tecnicatura Universitaria en Programacion a Distancia (UTN)
Catedra: Organizacion Empresarial - Trabajo Practico Integrador

Fase 3 completa: maquina de estados + integracion con datos_negocio.xlsx
+ calculo de recargo + cantidad de unidades + persistencia de pedidos
+ manejo de camino infeliz.
"""

import unicodedata
import csv
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment
    XLSX_DISPONIBLE = True
except ImportError:
    XLSX_DISPONIBLE = False

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------
RECARGO_TARJETA  = 0.10   # 10 % de recargo por pago con tarjeta
MAX_REINTENTOS   = 3      # intentos antes de derivar con asesor
MAX_CANTIDAD     = 99     # limite razonable de unidades por pedido

BASE_DIR   = Path(__file__).parent
RUTA_XLSX  = BASE_DIR / "data" / "datos_negocio.xlsx"
RUTA_CSV   = BASE_DIR / "data" / "pedidos.csv"

HEADERS_CSV = [
    "ID_Pedido", "Fecha_Hora", "Telefono", "Codigo_Repuesto",
    "Nombre_Repuesto", "Cantidad", "Precio_Base_Unit",
    "Metodo_Pago", "Recargo_Pct", "Precio_Unit_Final", "Total",
    "Envio_Retiro", "Estado"
]

METODOS_PAGO_VALIDOS   = {"efectivo", "transferencia", "tarjeta"}
OPCIONES_ENVIO_VALIDAS = {"envio", "retiro", "1", "2"}


# ---------------------------------------------------------------------------
# Utilidades de texto
# ---------------------------------------------------------------------------
SALUDOS_GENERICOS = {
    "hola", "buenas", "buen dia", "buenos dias",
    "buenas tardes", "buenas noches", "hello", "hi", "ola",
}

# Palabras que se ignoran al buscar productos
STOPWORDS = {
    "para", "mi", "un", "una", "el", "la", "los", "las",
    "busco", "busca", "buscando", "quiero", "necesito", "hay",
    "estoy", "tengo", "me", "te", "al", "en", "con", "por",
    "que", "como", "cual", "su", "si", "no", "del",
}

# Prefijos de saludo (ordenados de mas largo a mas corto)
PREFIJOS_SALUDO = [
    "buenas noches", "buenas tardes", "buenos dias", "buen dia",
    "buenas", "hola", "ola", "hello", "hi",
]


def normalizar(texto: str) -> str:
    texto = texto.strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto


def es_saludo_generico(mensaje: str) -> bool:
    return normalizar(mensaje) == "" or normalizar(mensaje) in SALUDOS_GENERICOS


def extraer_consulta(mensaje: str) -> str:
    """Quita el saludo del inicio y devuelve solo la parte de consulta."""
    norm = normalizar(mensaje)
    for prefijo in PREFIJOS_SALUDO:
        if norm.startswith(prefijo):
            resto = norm[len(prefijo):].strip().lstrip(",").lstrip(".").lstrip("!").strip()
            return resto
    return norm


# ---------------------------------------------------------------------------
# Capa de datos
# ---------------------------------------------------------------------------
def cargar_catalogo() -> list:
    if not XLSX_DISPONIBLE or not RUTA_XLSX.exists():
        return [
            {"codigo": "REP001", "nombre": "Pastillas de freno delanteras",
             "marca": "Chevrolet",  "modelo": "Corsa",  "precio": 12500, "stock": 8},
            {"codigo": "REP002", "nombre": "Filtro de aceite",
             "marca": "Chevrolet",  "modelo": "Corsa",  "precio":  2800, "stock": 15},
            {"codigo": "REP003", "nombre": "Amortiguador delantero",
             "marca": "Volkswagen", "modelo": "Gol",    "precio": 22000, "stock": 4},
            {"codigo": "REP010", "nombre": "Aceite de motor 5W40 (4L)",
             "marca": "Universal",  "modelo": "Todos",  "precio":  4800, "stock": 30},
            {"codigo": "REP011", "nombre": "Bateria 12V 65Ah",
             "marca": "Universal",  "modelo": "Todos",  "precio": 32000, "stock": 7},
            {"codigo": "REP014", "nombre": "Radiador",
             "marca": "Peugeot",    "modelo": "207",    "precio": 35000, "stock": 0},
        ]
    wb  = openpyxl.load_workbook(RUTA_XLSX, data_only=True)
    ws  = wb["Repuestos"]
    cat = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo, nombre, marca, modelo, anio, precio, stock = row
        if codigo:
            cat.append({
                "codigo":  str(codigo).strip(),
                "nombre":  str(nombre).strip(),
                "marca":   str(marca).strip(),
                "modelo":  str(modelo).strip(),
                "precio":  float(precio or 0),
                "stock":   int(stock or 0),
            })
    return cat


def buscar_repuesto(consulta: str, catalogo: list):
    consulta_n = normalizar(consulta)
    # Busqueda exacta por codigo
    for rep in catalogo:
        if normalizar(rep["codigo"]) == consulta_n:
            return rep
    # Busqueda por palabras clave, ignorando stopwords
    palabras = [p for p in consulta_n.split() if p not in STOPWORDS]
    if not palabras:
        return None
    mejores = []
    for rep in catalogo:
        texto_rep = normalizar(f"{rep['nombre']} {rep['marca']} {rep['modelo']}")
        if all(p in texto_rep for p in palabras):
            mejores.append(rep)
    if len(mejores) == 1:
        return mejores[0]
    if len(mejores) > 1:
        return sorted(mejores, key=lambda r: r["stock"], reverse=True)[0]
    return None


def buscar_todos(consulta: str, catalogo: list) -> list:
    """Devuelve TODOS los productos que coinciden con la consulta."""
    consulta_n = normalizar(consulta)
    # Busqueda exacta por codigo -> un solo resultado
    for rep in catalogo:
        if normalizar(rep["codigo"]) == consulta_n:
            return [rep]
    palabras = [p for p in consulta_n.split() if p not in STOPWORDS]
    if not palabras:
        return []
    return [
        rep for rep in catalogo
        if all(p in normalizar(f"{rep['nombre']} {rep['marca']} {rep['modelo']}") for p in palabras)
    ]


def guardar_pedido(sesion: dict):
    ahora    = datetime.now()
    id_ped   = ahora.strftime("PED%Y%m%d%H%M%S")
    rep      = sesion["repuesto_encontrado"]
    cantidad = sesion.get("cantidad", 1)
    recargo  = RECARGO_TARJETA if sesion["metodo_pago"] == "tarjeta" else 0.0
    precio_u = round(rep["precio"] * (1 + recargo), 2)
    total    = round(precio_u * cantidad, 2)

    fila = [
        id_ped,
        ahora.strftime("%Y-%m-%d %H:%M:%S"),
        sesion.get("telefono", "N/A"),
        rep["codigo"],
        rep["nombre"],
        cantidad,
        rep["precio"],
        sesion["metodo_pago"],
        f"{int(recargo*100)}%",
        precio_u,
        total,
        sesion["envio_o_retiro"],
        "CONFIRMADO",
    ]
    if XLSX_DISPONIBLE and RUTA_XLSX.exists():
        wb = openpyxl.load_workbook(RUTA_XLSX)
        ws = wb["Pedidos"]
        ws.append(fila)
        for col in range(1, len(fila) + 1):
            ws.cell(row=ws.max_row, column=col).alignment = Alignment(horizontal="center")
        wb.save(RUTA_XLSX)

    # Guardar tambien en CSV (visible en VSCode con Rainbow CSV)
    escribir_header = not RUTA_CSV.exists()
    with open(RUTA_CSV, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if escribir_header:
            writer.writerow(HEADERS_CSV)
        writer.writerow(fila)

    return id_ped, precio_u, total


# ---------------------------------------------------------------------------
# Estados
# ---------------------------------------------------------------------------
INICIO              = "INICIO"
ESPERANDO_REPUESTO  = "ESPERANDO_REPUESTO"
ESPERANDO_SELECCION = "ESPERANDO_SELECCION"
ESPERANDO_CANTIDAD  = "ESPERANDO_CANTIDAD"
ESPERANDO_PAGO      = "ESPERANDO_PAGO"
ESPERANDO_ENVIO     = "ESPERANDO_ENVIO"
PEDIDO_REGISTRADO   = "PEDIDO_REGISTRADO"
ESPERANDO_CONTINUAR = "ESPERANDO_CONTINUAR"
FIN_SIN_STOCK       = "FIN_SIN_STOCK"
FIN_DERIVADO        = "FIN_DERIVADO"


def crear_sesion(telefono: str = "N/A") -> dict:
    return {
        "estado":              INICIO,
        "telefono":            telefono,
        "repuesto_consultado": None,
        "repuesto_encontrado": None,
        "candidatos":          [],
        "cantidad":            1,
        "metodo_pago":         None,
        "envio_o_retiro":      None,
        "reintentos_pago":     0,
        "reintentos_envio":    0,
        "reintentos_cantidad": 0,
    }


# ---------------------------------------------------------------------------
# Manejadores
# ---------------------------------------------------------------------------
CATALOGO = cargar_catalogo()

BIENVENIDA = (
    "Hola! Bienvenido a *Tu Repuesto Cordoba*.\n"
    "Decime marca, modelo y anio del vehiculo, o el codigo "
    "del repuesto que buscas (ej: 'pastillas Corsa' o 'REP001')."
)


def _buscar_y_responder(sesion: dict, consulta: str):
    sesion["repuesto_consultado"] = consulta
    candidatos = buscar_todos(consulta, CATALOGO)

    if not candidatos:
        return (
            f"No encontre '{consulta}' en nuestro catalogo.\n"
            "Podes probar con otra descripcion, o escribi *asesor* para hablar con una persona."
        ), FIN_SIN_STOCK

    # Multiples resultados: pedir al cliente que elija
    if len(candidatos) > 1:
        sesion["candidatos"] = candidatos
        lineas = ["Encontre varias opciones. Cual es la que buscas?\n"]
        for i, r in enumerate(candidatos, 1):
            stock_txt = f"{r['stock']} disponibles" if r["stock"] > 0 else "sin stock"
            lineas.append(f"  *{i}.* {r['nombre']} ({r['marca']} {r['modelo']}) - {stock_txt}")
        lineas.append("\nEscribi el numero de la opcion.")
        return "\n".join(lineas), ESPERANDO_SELECCION

    rep = candidatos[0]
    if rep["stock"] == 0:
        sesion["repuesto_encontrado"] = rep
        return (
            f"Lo siento, *{rep['nombre']}* ({rep['marca']} {rep['modelo']}) "
            "esta temporalmente sin stock.\n"
            "Podes dejarnos tu numero y te avisamos cuando llegue, o escribi *asesor* para otras opciones."
        ), FIN_SIN_STOCK

    sesion["repuesto_encontrado"] = rep
    precio_fmt = f"${rep['precio']:,.0f}".replace(",", ".")
    return (
        f"Tenemos *{rep['nombre']}* ({rep['marca']} {rep['modelo']}) "
        f"en stock ({rep['stock']} disponibles).\n"
        f"Precio unitario: *{precio_fmt}*\n\n"
        "Cuantas unidades necesitas?"
    ), ESPERANDO_CANTIDAD


def manejar_inicio(sesion: dict, mensaje: str):
    # Saludo generico puro (sin consulta adicional)
    if es_saludo_generico(mensaje):
        return BIENVENIDA, ESPERANDO_REPUESTO

    # Extraer la parte de consulta quitando el saludo del inicio
    # Ej: "Hola estoy buscando filtro de aire" -> "estoy buscando filtro de aire"
    consulta = extraer_consulta(mensaje)
    if not consulta:
        return BIENVENIDA, ESPERANDO_REPUESTO

    # Intentar buscar directamente (las stopwords se ignoran en buscar_repuesto)
    rep = buscar_repuesto(consulta, CATALOGO)
    if rep is not None:
        encabezado = "Hola! Bienvenido a *Tu Repuesto Cordoba*.\n"
        respuesta, nuevo_estado = _buscar_y_responder(sesion, consulta)
        return encabezado + respuesta, nuevo_estado

    # No se encontro: saluda y pide mas datos
    sesion["repuesto_consultado"] = consulta
    return (
        "Hola! Bienvenido a *Tu Repuesto Cordoba*.\n"
        f"No encontre '{consulta}' en nuestro catalogo. "
        "Podes especificar la marca y modelo, o el codigo del repuesto "
        "(ej: 'pastillas Corsa' o 'REP001')."
    ), ESPERANDO_REPUESTO


def manejar_esperando_repuesto(sesion: dict, mensaje: str):
    texto = normalizar(mensaje)
    if texto == "":
        return (
            "No entendi tu consulta. "
            "Escribi la marca, modelo o codigo del repuesto que buscas."
        ), ESPERANDO_REPUESTO
    if texto == "asesor":
        return (
            "Te derivo con un asesor. En breve alguien se va a comunicar con vos. Hasta luego!"
        ), FIN_DERIVADO
    return _buscar_y_responder(sesion, mensaje)


def manejar_esperando_seleccion(sesion: dict, mensaje: str):
    texto = mensaje.strip()
    candidatos = sesion.get("candidatos", [])
    try:
        idx = int(texto) - 1
        if idx < 0 or idx >= len(candidatos):
            raise ValueError
    except ValueError:
        opciones = ", ".join(str(i+1) for i in range(len(candidatos)))
        return (
            f"Opcion no valida. Escribi un numero entre {opciones}."
        ), ESPERANDO_SELECCION

    rep = candidatos[idx]
    sesion["candidatos"] = []
    if rep["stock"] == 0:
        sesion["repuesto_encontrado"] = rep
        return (
            f"Lo siento, *{rep['nombre']}* ({rep['marca']} {rep['modelo']}) "
            "esta temporalmente sin stock.\n"
            "Podes dejarnos tu numero y te avisamos cuando llegue, o escribi *asesor* para otras opciones."
        ), FIN_SIN_STOCK

    sesion["repuesto_encontrado"] = rep
    precio_fmt = f"${rep['precio']:,.0f}".replace(",", ".")
    return (
        f"Perfecto! *{rep['nombre']}* ({rep['marca']} {rep['modelo']}) "
        f"en stock ({rep['stock']} disponibles).\n"
        f"Precio unitario: *{precio_fmt}*\n\n"
        "Cuantas unidades necesitas?"
    ), ESPERANDO_CANTIDAD


def manejar_esperando_cantidad(sesion: dict, mensaje: str):
    texto = mensaje.strip()
    rep   = sesion["repuesto_encontrado"]

    try:
        cantidad = int(texto)
        if cantidad <= 0:
            raise ValueError
    except ValueError:
        sesion["reintentos_cantidad"] += 1
        if sesion["reintentos_cantidad"] >= MAX_REINTENTOS:
            return (
                "No pude entender la cantidad despues de varios intentos. Te derivo con un asesor."
            ), FIN_DERIVADO
        return (
            f"'{texto}' no es una cantidad valida. "
            "Escribi un numero entero mayor a cero (ej: 1, 2, 3...)."
        ), ESPERANDO_CANTIDAD

    if cantidad > rep["stock"]:
        sesion["reintentos_cantidad"] += 1
        if sesion["reintentos_cantidad"] >= MAX_REINTENTOS:
            return "Superaste el limite de intentos. Te derivo con un asesor.", FIN_DERIVADO
        return (
            f"Solo tenemos *{rep['stock']}* unidades disponibles de {rep['nombre']}. "
            f"Cuantas necesitas (maximo {rep['stock']})?"
        ), ESPERANDO_CANTIDAD

    sesion["cantidad"]            = cantidad
    sesion["reintentos_cantidad"] = 0

    precio_fmt = f"${rep['precio']:,.0f}".replace(",", ".")
    if cantidad > 1:
        subtotal = f"${rep['precio'] * cantidad:,.0f}".replace(",", ".")
        detalle  = f"{cantidad} x {precio_fmt} = {subtotal}"
    else:
        detalle = precio_fmt

    return (
        f"Perfecto, *{cantidad} unidad{'es' if cantidad > 1 else ''}* de {rep['nombre']} ({detalle}).\n\n"
        "Como preferis pagar?\n"
        "  *efectivo* | *transferencia* | *tarjeta* (+ 10% recargo)"
    ), ESPERANDO_PAGO


def manejar_esperando_pago(sesion: dict, mensaje: str):
    texto = normalizar(mensaje)

    if texto not in METODOS_PAGO_VALIDOS:
        sesion["reintentos_pago"] += 1
        if sesion["reintentos_pago"] >= MAX_REINTENTOS:
            return (
                "Hubo demasiados intentos sin una opcion valida. Te derivo con un asesor."
            ), FIN_DERIVADO
        return (
            f"Opcion no reconocida: '{mensaje}'.\n"
            "Por favor elegi: *efectivo*, *transferencia* o *tarjeta*."
        ), ESPERANDO_PAGO

    sesion["metodo_pago"] = texto
    rep      = sesion["repuesto_encontrado"]
    cantidad = sesion.get("cantidad", 1)
    recargo  = RECARGO_TARJETA if texto == "tarjeta" else 0.0
    precio_u = round(rep["precio"] * (1 + recargo), 2)
    total    = round(precio_u * cantidad, 2)
    total_fmt = f"${total:,.0f}".replace(",", ".")

    detalle = ""
    if recargo:
        base_fmt = f"${rep['precio'] * cantidad:,.0f}".replace(",", ".")
        detalle  = f" ({base_fmt} + 10% tarjeta = {total_fmt})"
    elif cantidad > 1:
        detalle = f" (total {cantidad} unidades: {total_fmt})"

    return (
        f"Perfecto, pago con *{texto}*{detalle}.\n\n"
        "Como queres recibir el repuesto?\n"
        "  *1 - Envio* a domicilio | *2 - Retiro* en local (Av. Colon 1234, Cordoba)"
    ), ESPERANDO_ENVIO


def manejar_esperando_envio(sesion: dict, mensaje: str):
    texto = normalizar(mensaje)

    if texto in {"1", "envio", "enviar", "lo mandan", "lo envian", "domicilio"}:
        sesion["envio_o_retiro"] = "envio"
    elif texto in {"2", "retiro", "retira", "voy a buscar", "paso a buscar", "local"}:
        sesion["envio_o_retiro"] = "retiro"
    else:
        sesion["reintentos_envio"] += 1
        if sesion["reintentos_envio"] >= MAX_REINTENTOS:
            return (
                "No pude entender tu respuesta despues de varios intentos. Te derivo con un asesor."
            ), FIN_DERIVADO
        return (
            f"No entendi '{mensaje}'.\n"
            "Escribi *1* para envio a domicilio o *2* para retiro en local."
        ), ESPERANDO_ENVIO

    id_ped, precio_u, total = guardar_pedido(sesion)
    precio_u_fmt = f"${precio_u:,.0f}".replace(",", ".")
    total_fmt    = f"${total:,.0f}".replace(",", ".")
    modo = "Envio a domicilio" if sesion["envio_o_retiro"] == "envio" else "Retiro en local"
    rep      = sesion["repuesto_encontrado"]
    cantidad = sesion.get("cantidad", 1)

    return (
        "Tu pedido quedo *confirmado*!\n\n"
        "Resumen:\n"
        f"  Repuesto  : {rep['nombre']}\n"
        f"  Cantidad  : {cantidad} unidad{'es' if cantidad > 1 else ''}\n"
        f"  Precio u. : {precio_u_fmt}\n"
        f"  Total     : *{total_fmt}*\n"
        f"  Pago      : {sesion['metodo_pago']}\n"
        f"  Entrega   : {modo}\n"
        f"  N pedido  : {id_ped}\n\n"
        "Queres consultar otro repuesto? (*si* / *no*)"
    ), ESPERANDO_CONTINUAR


def manejar_esperando_continuar(sesion: dict, mensaje: str):
    texto = normalizar(mensaje)
    if texto in {"si", "si", "s", "yes", "dale", "quiero", "otra consulta"}:
        tel = sesion.get("telefono", "N/A")
        sesion.update(crear_sesion(tel))
        return BIENVENIDA, ESPERANDO_REPUESTO
    if texto in {"no", "n", "no gracias", "listo", "nope"}:
        return (
            "Perfecto! Gracias por elegirnos. Hasta la proxima!"
        ), PEDIDO_REGISTRADO
    return (
        "No entendi tu respuesta. Escribi *si* para hacer otra consulta o *no* para terminar."
    ), ESPERANDO_CONTINUAR


def manejar_estado_terminal(sesion: dict, mensaje: str):
    texto = normalizar(mensaje)

    if texto in {"hola", "nueva consulta", "reiniciar", "volver"}:
        tel = sesion.get("telefono", "N/A")
        sesion.update(crear_sesion(tel))
        return BIENVENIDA, ESPERANDO_REPUESTO

    if sesion["estado"] == FIN_SIN_STOCK and texto not in {"", "asesor"}:
        return _buscar_y_responder(sesion, mensaje)

    if sesion["estado"] == FIN_SIN_STOCK:
        return (
            "Podes buscar otro repuesto escribiendo la marca y modelo, "
            "o escribi *hola* para empezar de nuevo."
        ), FIN_SIN_STOCK

    return (
        "Tu consulta ya fue procesada. Si queres hacer otra, "
        "escribi *hola* o *nueva consulta*."
    ), sesion["estado"]


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------
MANEJADORES = {
    INICIO:               manejar_inicio,
    ESPERANDO_REPUESTO:   manejar_esperando_repuesto,
    ESPERANDO_SELECCION:  manejar_esperando_seleccion,
    ESPERANDO_CANTIDAD:   manejar_esperando_cantidad,
    ESPERANDO_PAGO:      manejar_esperando_pago,
    ESPERANDO_ENVIO:     manejar_esperando_envio,
    ESPERANDO_CONTINUAR: manejar_esperando_continuar,
    PEDIDO_REGISTRADO:   manejar_estado_terminal,
    FIN_SIN_STOCK:       manejar_estado_terminal,
    FIN_DERIVADO:        manejar_estado_terminal,
}


def procesar_mensaje(sesion: dict, mensaje: str) -> str:
    manejador = MANEJADORES.get(sesion["estado"], manejar_estado_terminal)
    respuesta, nuevo_estado = manejador(sesion, mensaje)
    sesion["estado"] = nuevo_estado
    return respuesta


# ---------------------------------------------------------------------------
# Simulacion por consola
# ---------------------------------------------------------------------------
# Estados que indican que la conversacion termino
ESTADOS_FINALES = {PEDIDO_REGISTRADO, FIN_DERIVADO}

def simular_conversacion():
    (BASE_DIR / "data").mkdir(exist_ok=True)
    sesiones      = {}
    telefono_demo = "+549351000000"
    sesiones[telefono_demo] = crear_sesion(telefono_demo)

    print("=" * 55)
    print("  Simulacion de bot - Tu Repuesto Cordoba")
    print("  (escribi 'salir' para terminar)")
    print("=" * 55)
    print()

    while True:
        try:
            msg = input("Cliente: ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\n[Simulacion interrumpida]")
            break
        if msg.lower() == "salir":
            print("Bot: Hasta luego!")
            break
        sesion = sesiones[telefono_demo]
        respuesta = procesar_mensaje(sesion, msg)
        print(f"\nBot: {respuesta}\n")
        # Terminar la simulacion si la conversacion llego a un estado final
        if sesion["estado"] in ESTADOS_FINALES:
            print("[Conversacion finalizada. Ejecuta el script nuevamente para iniciar otra.]")
            break


if __name__ == "__main__":
    simular_conversacion()